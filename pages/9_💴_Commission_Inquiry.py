import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import re

# 1. PAGE SETUP
st.set_page_config(page_title="Commission Inquiry", page_icon="💴", layout="wide")

# Custom CSS for consistent "Nebula" theme
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
        color: #e2e8f0;
    }

    .stApp {
        background-color: #030712;
        background-image: 
            radial-gradient(at 0% 0%, rgba(56, 189, 248, 0.1) 0px, transparent 50%),
            radial-gradient(at 100% 0%, rgba(139, 92, 246, 0.1) 0px, transparent 50%);
        background-attachment: fixed;
    }

    .css-1d391kg {
        background: rgba(15, 23, 42, 0.6);
        backdrop-filter: blur(10px);
        border-radius: 16px;
        border: 1px solid rgba(255, 255, 255, 0.05);
        padding: 2rem;
    }

    h1, h2, h3 {
        color: #fff;
        text-shadow: 0 0 10px rgba(255, 255, 255, 0.1);
    }
</style>
""", unsafe_allow_html=True)

st.title("💴 Commission Inquiry")
st.write("Scan 20 schedule installments for specific months and detect red-marked records.")

def get_schedule_cols(df_cols):
    schedule_map = {}
    for i in range(1, 21):
        month_col = None
        comm_col = None
        for c in df_cols:
            if re.search(rf"^{i}[-\s]*Schedule\s*Month", str(c), re.I):
                month_col = c
            if re.search(rf"^{i}[-\s]*Schedule\s*Commission", str(c), re.I):
                comm_col = c
        if month_col and comm_col:
            schedule_map[i] = {"month": month_col, "commission": comm_col}
    return schedule_map

@st.cache_data
def process_file(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    if "Master File" not in wb.sheetnames:
        return None, "Sheet 'Master File' not found."
    
    sheet = wb["Master File"]
    header_row_idx = 2
    cols = []
    for col in range(1, sheet.max_column + 1):
        cols.append(sheet.cell(row=header_row_idx, column=col).value)
    
    data = []
    base_cols = {
        "Agentcis ID": 1, "App ID": 2, "Student Name": 3,
        "Email": 6, "COE": 11, "COE Type": 12, "Provider Name": 13
    }
    
    col_names = [str(c) for c in cols if c]
    schedule_mapping = get_schedule_cols(col_names)
    
    idx_map = {}
    for i, meta in schedule_mapping.items():
        m_name, c_name = meta["month"], meta["commission"]
        m_idx = next((idx for idx, name in enumerate(cols, 1) if name == m_name), None)
        c_idx = next((idx for idx, name in enumerate(cols, 1) if name == c_name), None)
        idx_map[i] = {"month": m_idx, "commission": c_idx}

    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
        # Check student name (Col C/3) as anchor
        if not sheet.cell(row=row_idx, column=3).value: continue
            
        row_data = {}
        cell_app_id = sheet.cell(row=row_idx, column=2)
        is_red = cell_app_id.fill and cell_app_id.fill.start_color.index == 'FFC00000'
        row_data["red marked"] = "ss" if is_red else ""
        
        for label, col_idx in base_cols.items():
            row_data[label] = sheet.cell(row=row_idx, column=col_idx).value
            
        row_data["installments"] = []
        for i, indices in idx_map.items():
            if indices["month"] and indices["commission"]:
                month_val = sheet.cell(row=row_idx, column=indices["month"]).value
                comm_val = sheet.cell(row=row_idx, column=indices["commission"]).value
                if month_val:
                    row_data["installments"].append({
                        "schedule_no": i,
                        "month": str(month_val).strip(),
                        "commission": comm_val
                    })
        data.append(row_data)
    return data, None

uploaded_file = st.file_uploader("Upload Adelaide Sales Tracker (.xlsx)", type=["xlsx"])

if uploaded_file:
    with st.spinner("Analyzing commissions..."):
        file_bytes = uploaded_file.read()
        data, error = process_file(file_bytes)
        
    if error: st.error(error)
    else:
        st.success(f"Processed {len(data)} records.")
        
        all_months = set()
        for row in data:
            for inst in row.get("installments", []):
                if inst["month"] and inst["month"] != "None":
                    all_months.add(inst["month"])
        
        selected_month = st.selectbox("Select Target Month", sorted(list(all_months)))
        
        if selected_month:
            filtered = []
            for row in data:
                for inst in row.get("installments", []):
                    if inst["month"].strip() == selected_month.strip():
                        filtered.append({
                            "Red Marked": row["red marked"],
                            "Agentcis ID": row["Agentcis ID"],
                            "App ID": row["App ID"],
                            "Student": row["Student Name"],
                            "Email": row["Email"],
                            "COE": row["COE"],
                            "COE Type": row["COE Type"],
                            "Provider": row["Provider Name"],
                            "Schedule": inst["schedule_no"],
                            "Commission": inst["commission"]
                        })
            
            if filtered:
                st.write(f"### Results for {selected_month} ({len(filtered)} items)")
                st.dataframe(pd.DataFrame(filtered), use_container_width=True)
                csv = pd.DataFrame(filtered).to_csv(index=False).encode('utf-8')
                st.download_button("📥 Export to CSV", csv, f"commissions_{selected_month}.csv", "text/csv")
            else:
                st.info("No records found for this month.")
